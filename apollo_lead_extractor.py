"""
Apollo.io People Enrichment API 기반 B2B 담당자 추출 시스템
============================================================
People Enrichment API (POST /api/v1/people/match)를 사용하여
담당자 정보를 enrichment하고, 이메일/상세정보를 확보합니다.

입력 방식:
  A) CSV/Excel 파일: 이름 + 회사(도메인) 리스트 업로드
  B) 직접 입력: 파이썬 리스트로 담당자 정보 전달
  C) LinkedIn URL 기반: LinkedIn 프로필 URL로 enrichment

사용법:
1. Apollo.io에서 API 키 발급
2. .env 파일에 APOLLO_API_KEY=your_key 설정
3. 입력 파일(CSV/Excel) 준비 또는 코드에서 직접 입력
4. python apollo_lead_extractor.py 실행

주의: Enrichment API는 크레딧을 소모합니다!
"""

import requests
import json
import csv
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

# ============================================================
# 설정
# ============================================================

APOLLO_API_KEY = ""  # 직접 입력 또는 .env 사용


def load_api_key():
    global APOLLO_API_KEY
    if APOLLO_API_KEY:
        return APOLLO_API_KEY
    env_path = Path(".env")
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            line = line.strip()
            if line.startswith("APOLLO_API_KEY="):
                APOLLO_API_KEY = line.split("=", 1)[1].strip().strip('"').strip("'")
                return APOLLO_API_KEY
    raise ValueError(
        "API 키가 설정되지 않았습니다.\n"
        "1) .env 파일에 APOLLO_API_KEY=your_key 추가\n"
        "2) 또는 스크립트 상단 APOLLO_API_KEY 변수에 직접 입력"
    )


# ============================================================
# 입력 데이터 예시 (사용자가 수정)
# ============================================================

ENRICHMENT_TARGETS = [
    # 최소: first_name + last_name + domain (또는 organization_name)
    # 정보가 많을수록 매칭 정확도 상승
    {
        "first_name": "Tim",
        "last_name": "Zheng",
        "domain": "apollo.io",
    },
    {
        "first_name": "Satya",
        "last_name": "Nadella",
        "organization_name": "Microsoft",
    },
    # LinkedIn URL만으로도 가능
    {
        "linkedin_url": "https://www.linkedin.com/in/example",
    },
    # 이메일만으로도 가능
    {
        "email": "example@company.com",
    },
]

# Enrichment 옵션
ENRICHMENT_OPTIONS = {
    "reveal_personal_emails": False,  # True면 개인 이메일 공개 (크레딧 추가 소모)
    "reveal_phone_number": False,     # True면 전화번호 공개 (webhook_url 필수)
    "run_waterfall_email": False,     # True면 워터폴 이메일 enrichment
    "run_waterfall_phone": False,     # True면 워터폴 전화번호 enrichment
}


# ============================================================
# Apollo People Enrichment API 클라이언트
# ============================================================

class ApolloEnrichmentClient:
    """Apollo.io People Enrichment API 클라이언트"""

    BASE_URL = "https://api.apollo.io/api/v1"
    PARAM_KEYS = [
        "first_name", "last_name", "name", "email", "hashed_email",
        "organization_name", "domain", "id", "linkedin_url",
    ]

    def __init__(self, api_key: str):
        self.api_key = api_key
        self.session = requests.Session()
        self.session.headers.update({
            "Content-Type": "application/json",
            "Cache-Control": "no-cache",
            "x-api-key": api_key,
        })
        self._request_count = 0
        self._credits_used = 0

    def enrich_person(self, params: dict, options: dict = None) -> dict:
        """
        단일 인물 Enrichment (POST /api/v1/people/match)

        Args:
            params: 검색 파라미터
                - first_name + last_name + domain (추천 조합)
                - email
                - linkedin_url
                - id (Apollo ID)
                - name + organization_name
            options: reveal_personal_emails, reveal_phone_number 등
        """
        options = options or ENRICHMENT_OPTIONS
        query_params = {}

        for key in self.PARAM_KEYS:
            if params.get(key):
                query_params[key] = params[key]

        query_params["reveal_personal_emails"] = options.get("reveal_personal_emails", False)
        query_params["reveal_phone_number"] = options.get("reveal_phone_number", False)
        query_params["run_waterfall_email"] = options.get("run_waterfall_email", False)
        query_params["run_waterfall_phone"] = options.get("run_waterfall_phone", False)

        if options.get("webhook_url"):
            query_params["webhook_url"] = options["webhook_url"]

        return self._request(query_params)

    def enrich_bulk(self, people_list: list[dict], options: dict = None) -> list[dict]:
        """
        벌크 Enrichment (POST /api/v1/people/bulk_match, 최대 10명씩)
        """
        options = options or ENRICHMENT_OPTIONS
        results = []
        total = len(people_list)
        batch_size = 10

        for batch_start in range(0, total, batch_size):
            batch = people_list[batch_start:batch_start + batch_size]
            batch_end = min(batch_start + batch_size, total)
            print(f"\n📦 배치 {batch_start + 1}-{batch_end} / {total}")

            details = []
            for person_params in batch:
                detail = {k: person_params[k] for k in self.PARAM_KEYS if person_params.get(k)}
                details.append(detail)

            payload = {
                "details": details,
                "reveal_personal_emails": options.get("reveal_personal_emails", False),
                "reveal_phone_number": options.get("reveal_phone_number", False),
            }

            try:
                resp = self.session.post(f"{self.BASE_URL}/people/bulk_match", json=payload, timeout=30)

                if resp.status_code == 429:
                    print("  ⏳ 레이트 리밋. 60초 대기...")
                    time.sleep(60)
                    resp = self.session.post(f"{self.BASE_URL}/people/bulk_match", json=payload, timeout=30)

                if resp.status_code == 200:
                    data = resp.json()
                    matches = data.get("matches", []) or []
                    for i, match in enumerate(matches):
                        src = batch[i]
                        identifier = src.get("email") or src.get("domain") or src.get("linkedin_url") or ""
                        name_str = f"{src.get('first_name', '')} {src.get('last_name', '')}".strip()
                        if match:
                            results.append(match)
                            self._credits_used += 1
                            print(f"  ✅ {name_str} ({identifier})")
                        else:
                            print(f"  ⚠️ 매칭 실패: {name_str} ({identifier})")
                else:
                    print(f"  ❌ API 오류 ({resp.status_code}): {resp.text[:200]}")

            except requests.exceptions.RequestException as e:
                print(f"  ❌ 요청 실패: {e}")

            if batch_start + batch_size < total:
                time.sleep(1)

        return results

    def _request(self, query_params: dict) -> dict:
        """단일 Enrichment 요청"""
        self._request_count += 1
        if self._request_count % 10 == 0:
            time.sleep(1)

        try:
            resp = self.session.post(f"{self.BASE_URL}/people/match", params=query_params, timeout=30)

            if resp.status_code == 429:
                print("  ⏳ 레이트 리밋. 60초 대기...")
                time.sleep(60)
                return self._request(query_params)

            if resp.status_code == 200:
                data = resp.json()
                if data.get("person"):
                    self._credits_used += 1
                return data
            else:
                print(f"  ⚠️ API ({resp.status_code}): {resp.text[:200]}")
                return {}
        except requests.exceptions.RequestException as e:
            print(f"  ❌ 요청 실패: {e}")
            return {}

    def get_credits_used(self) -> int:
        return self._credits_used


# ============================================================
# 데이터 처리 & 정제
# ============================================================

class LeadProcessor:
    """Enrichment 응답 데이터 처리 및 정제"""

    def __init__(self):
        self._seen_emails = set()
        self._seen_ids = set()

    def parse_enrichment(self, raw: dict, source_info: dict = None) -> Optional[dict]:
        """People Enrichment API 응답 파싱"""
        if not raw:
            return None

        person_id = raw.get("id", "")
        email = raw.get("email", "") or ""

        if person_id and person_id in self._seen_ids:
            return None
        if person_id:
            self._seen_ids.add(person_id)
        if email and email in self._seen_emails:
            return None
        if email:
            self._seen_emails.add(email)

        org = raw.get("organization", {}) or {}
        current_emp = self._get_current_employment(raw.get("employment_history", []))

        return {
            "이름": raw.get("name", "") or f"{raw.get('first_name', '')} {raw.get('last_name', '')}".strip(),
            "직함": raw.get("title", "") or (current_emp.get("title", "") if current_emp else ""),
            "직급": raw.get("seniority", ""),
            "부서": ", ".join(raw.get("departments", []) or []),
            "하위부서": ", ".join(raw.get("subdepartments", []) or []),
            "이메일": email,
            "이메일_상태": raw.get("email_status", "unknown"),
            "전화번호": self._get_phone(raw),
            "LinkedIn": raw.get("linkedin_url", ""),
            "Twitter": raw.get("twitter_url", ""),
            "회사명": org.get("name", ""),
            "회사_도메인": org.get("primary_domain", "") or org.get("website_url", ""),
            "회사_산업": org.get("industry", ""),
            "회사_규모": org.get("estimated_num_employees", ""),
            "회사_매출": org.get("annual_revenue_printed", "") or "",
            "회사_위치": self._get_org_location(org),
            "회사_LinkedIn": org.get("linkedin_url", ""),
            "회사_설명": (org.get("short_description", "") or "")[:200],
            "참여_가능성": "높음" if raw.get("is_likely_to_engage") else "보통",
            "Apollo_ID": person_id,
            "수집일시": datetime.now().strftime("%Y-%m-%d %H:%M"),
        }

    @staticmethod
    def _get_current_employment(history: list) -> Optional[dict]:
        if not history:
            return None
        for emp in history:
            if emp.get("current"):
                return emp
        return history[0] if history else None

    @staticmethod
    def _get_phone(raw: dict) -> str:
        phones = raw.get("phone_numbers", []) or []
        return phones[0].get("sanitized_number", "") or phones[0].get("raw_number", "") if phones else ""

    @staticmethod
    def _get_org_location(org: dict) -> str:
        parts = [org.get("city", ""), org.get("state", ""), org.get("country", "")]
        return ", ".join(p for p in parts if p)

    def validate_lead(self, lead: dict) -> dict:
        """리드 품질 검증 및 점수 부여"""
        score = 0
        issues = []

        if lead.get("이메일"):
            score += 30
            if lead.get("이메일_상태") == "verified":
                score += 20
            elif lead.get("이메일_상태") == "guessed":
                score += 5
                issues.append("이메일 추정값")
            else:
                issues.append(f"이메일 상태: {lead.get('이메일_상태', 'unknown')}")
        else:
            issues.append("이메일 없음")

        if lead.get("이름"): score += 10
        if lead.get("직함"): score += 10
        if lead.get("회사명"): score += 10
        if lead.get("LinkedIn"): score += 10
        if lead.get("전화번호"): score += 10

        lead["품질_점수"] = score
        lead["이슈"] = "; ".join(issues) if issues else "없음"
        return lead


# ============================================================
# 입력 파일 로더
# ============================================================

def load_targets_from_csv(filepath: str) -> list[dict]:
    """
    CSV에서 enrichment 대상 로드

    지원 컬럼: first_name, last_name, name, email, domain,
              organization_name, linkedin_url (한국어 매핑 지원)
    """
    column_map = {
        "이름": "name", "성": "last_name", "이름(영문)": "first_name",
        "first_name": "first_name", "last_name": "last_name", "name": "name",
        "이메일": "email", "email": "email",
        "회사": "organization_name", "회사명": "organization_name",
        "company": "organization_name", "organization_name": "organization_name",
        "도메인": "domain", "domain": "domain", "회사_도메인": "domain",
        "linkedin": "linkedin_url", "linkedin_url": "linkedin_url", "LinkedIn": "linkedin_url",
    }
    targets = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            target = {}
            for col, val in row.items():
                mapped = column_map.get(col.strip(), col.strip())
                if val and val.strip():
                    target[mapped] = val.strip()
            if target:
                targets.append(target)
    print(f"📂 {len(targets)}건 로드 (CSV: {filepath})")
    return targets


def load_targets_from_excel(filepath: str) -> list[dict]:
    """Excel에서 enrichment 대상 로드"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ openpyxl 필요: pip install openpyxl")
        return []

    column_map = {
        "이름": "name", "성": "last_name", "이름(영문)": "first_name",
        "first_name": "first_name", "last_name": "last_name", "name": "name",
        "이메일": "email", "email": "email",
        "회사": "organization_name", "회사명": "organization_name",
        "company": "organization_name", "organization_name": "organization_name",
        "도메인": "domain", "domain": "domain", "회사_도메인": "domain",
        "linkedin": "linkedin_url", "linkedin_url": "linkedin_url", "LinkedIn": "linkedin_url",
    }

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    targets = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        target = {}
        for i, value in enumerate(row):
            if i < len(headers) and value:
                mapped = column_map.get(headers[i], headers[i])
                target[mapped] = str(value).strip()
        if target:
            targets.append(target)
    print(f"📂 {len(targets)}건 로드 (Excel: {filepath})")
    return targets


# ============================================================
# 데이터 저장
# ============================================================

class LeadExporter:
    @staticmethod
    def to_csv(leads: list[dict], filepath: str):
        if not leads:
            return
        fp = Path(filepath)
        fp.parent.mkdir(parents=True, exist_ok=True)
        with open(fp, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=list(leads[0].keys()))
            w.writeheader()
            w.writerows(leads)
        print(f"✅ CSV 저장: {fp} ({len(leads)}건)")

    @staticmethod
    def to_excel(leads: list[dict], filepath: str):
        if not leads:
            return
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        fp = Path(filepath)
        fp.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "담당자 리스트"

        hfont = Font(bold=True, color="FFFFFF", size=11, name="Arial")
        hfill = PatternFill("solid", fgColor="2B5797")
        halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
        colors = {
            "high": PatternFill("solid", fgColor="C6EFCE"),
            "mid": PatternFill("solid", fgColor="FFEB9C"),
            "low": PatternFill("solid", fgColor="FFC7CE"),
        }

        headers = list(leads[0].keys())
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, border

        for ri, lead in enumerate(leads, 2):
            for ci, key in enumerate(headers, 1):
                val = lead.get(key, "")
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = border
                c.alignment = Alignment(vertical="center")
                if key == "품질_점수" and isinstance(val, (int, float)):
                    c.fill = colors["high"] if val >= 70 else colors["mid"] if val >= 40 else colors["low"]
                if key == "이메일_상태":
                    c.fill = colors["high"] if val == "verified" else colors["mid"] if val == "guessed" else colors["low"]
                if key == "참여_가능성":
                    c.fill = colors["high"] if val == "높음" else colors["mid"]

        widths = {"이름": 18, "직함": 25, "직급": 12, "부서": 15, "하위부서": 15,
                  "이메일": 30, "이메일_상태": 12, "전화번호": 18, "LinkedIn": 35,
                  "Twitter": 25, "회사명": 22, "회사_도메인": 22, "회사_산업": 18,
                  "회사_규모": 10, "회사_매출": 15, "회사_위치": 20, "회사_LinkedIn": 35,
                  "회사_설명": 30, "참여_가능성": 10, "Apollo_ID": 18, "수집일시": 16,
                  "품질_점수": 10, "이슈": 20}
        for ci, h in enumerate(headers, 1):
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = widths.get(h, 15)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        # 통계 시트
        ws2 = wb.create_sheet("통계 요약")
        stats = _compute_stats(leads)
        ws2["A1"] = "Enrichment 통계"
        ws2["A1"].font = Font(bold=True, size=14, name="Arial")
        rows = [("총 매칭", stats["total"]), ("이메일 확보", stats["with_email"]),
                ("이메일 verified", stats["verified"]), ("이메일 guessed", stats["guessed"]),
                ("평균 품질", f"{stats['avg_score']:.1f}"), ("참여가능성 높음", stats.get("likely_engage", 0)),
                ("", ""), ("직급별", "")]
        for k, v in stats.get("by_seniority", {}).items():
            rows.append((f"  {k}", v))
        rows += [("", ""), ("상위 회사", "")]
        for k, v in list(stats.get("by_company", {}).items())[:10]:
            rows.append((f"  {k}", v))
        for i, (l, v) in enumerate(rows, 3):
            ws2.cell(row=i, column=1, value=l)
            ws2.cell(row=i, column=2, value=v)
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 15

        wb.save(fp)
        print(f"✅ Excel 저장: {fp} ({len(leads)}건)")

    @staticmethod
    def to_json(leads: list[dict], filepath: str):
        fp = Path(filepath)
        fp.parent.mkdir(parents=True, exist_ok=True)
        with open(fp, "w", encoding="utf-8") as f:
            json.dump(leads, f, ensure_ascii=False, indent=2)
        print(f"✅ JSON 저장: {fp} ({len(leads)}건)")


def _compute_stats(leads):
    total = len(leads)
    with_email = sum(1 for l in leads if l.get("이메일"))
    verified = sum(1 for l in leads if l.get("이메일_상태") == "verified")
    guessed = sum(1 for l in leads if l.get("이메일_상태") == "guessed")
    likely = sum(1 for l in leads if l.get("참여_가능성") == "높음")
    scores = [l.get("품질_점수", 0) for l in leads]
    avg = sum(scores) / max(len(scores), 1)
    by_sen = {}
    for l in leads:
        s = l.get("직급", "기타") or "기타"
        by_sen[s] = by_sen.get(s, 0) + 1
    by_co = {}
    for l in leads:
        c = l.get("회사명", "기타") or "기타"
        by_co[c] = by_co.get(c, 0) + 1
    by_co = dict(sorted(by_co.items(), key=lambda x: x[1], reverse=True))
    return {"total": total, "with_email": with_email, "verified": verified,
            "guessed": guessed, "avg_score": avg, "likely_engage": likely,
            "by_seniority": by_sen, "by_company": by_co}


# ============================================================
# 메인 실행 함수들
# ============================================================

def run_enrichment(targets=None, options=None, use_bulk=True, output_prefix="enriched_leads"):
    """People Enrichment 메인 실행"""
    targets = targets or ENRICHMENT_TARGETS
    options = options or ENRICHMENT_OPTIONS

    print("=" * 60)
    print("🚀 Apollo People Enrichment 시작")
    print(f"   대상: {len(targets)}건")
    print(f"   모드: {'Bulk (10건 배치)' if use_bulk else '단건'}")
    print(f"   개인이메일: {options.get('reveal_personal_emails', False)}")
    print(f"   전화번호: {options.get('reveal_phone_number', False)}")
    print("=" * 60)

    api_key = load_api_key()
    client = ApolloEnrichmentClient(api_key)
    processor = LeadProcessor()
    all_leads = []

    if use_bulk and len(targets) > 1:
        results = client.enrich_bulk(targets, options)
        for raw in results:
            parsed = processor.parse_enrichment(raw)
            if parsed:
                all_leads.append(processor.validate_lead(parsed))
    else:
        for i, target in enumerate(targets, 1):
            ident = target.get("email") or target.get("domain") or target.get("linkedin_url") or ""
            name = f"{target.get('first_name', '')} {target.get('last_name', '')}".strip() or target.get("name", "")
            print(f"\n[{i}/{len(targets)}] {name} ({ident})")

            result = client.enrich_person(target, options)
            person = result.get("person")
            if person:
                parsed = processor.parse_enrichment(person, target)
                if parsed:
                    all_leads.append(processor.validate_lead(parsed))
                    print(f"  ✅ {parsed['이름']} | {parsed['직함']} | {parsed['이메일']}")
            else:
                print(f"  ⚠️ 매칭 실패")
            time.sleep(0.5)

    # 결과 요약
    print("\n" + "=" * 60)
    print("📊 Enrichment 결과")
    print("=" * 60)
    stats = _compute_stats(all_leads)
    print(f"  입력: {len(targets)}건 → 매칭: {stats['total']}건")
    print(f"  이메일 확보: {stats['with_email']}건 (verified: {stats['verified']})")
    print(f"  참여 가능성 높음: {stats.get('likely_engage', 0)}건")
    print(f"  평균 품질: {stats['avg_score']:.1f}/100")
    print(f"  크레딧 사용: ~{client.get_credits_used()}건")

    if not all_leads:
        print("\n⚠️ 매칭 실패. first_name + last_name + domain 조합을 추천합니다.")
        return []

    all_leads.sort(key=lambda x: x.get("품질_점수", 0), reverse=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    Path("output").mkdir(exist_ok=True)
    LeadExporter.to_excel(all_leads, f"output/{output_prefix}_{ts}.xlsx")
    LeadExporter.to_csv(all_leads, f"output/{output_prefix}_{ts}.csv")
    LeadExporter.to_json(all_leads, f"output/{output_prefix}_{ts}.json")
    print(f"\n✅ output/ 폴더에 저장 완료")
    return all_leads


def enrich_from_file(filepath, options=None, use_bulk=True, max_count=0):
    """파일(CSV/Excel)에서 대상 로드 → Enrichment"""
    if filepath.endswith(".csv"):
        targets = load_targets_from_csv(filepath)
    elif filepath.endswith((".xlsx", ".xls")):
        targets = load_targets_from_excel(filepath)
    else:
        print(f"❌ 지원하지 않는 형식: {filepath}")
        return []
    if max_count > 0:
        targets = targets[:max_count]
    return run_enrichment(targets, options, use_bulk)


def enrich_by_linkedin(urls, options=None):
    """LinkedIn URL 리스트로 Enrichment"""
    return run_enrichment([{"linkedin_url": u} for u in urls if u], options, True, "linkedin_enriched")


def enrich_by_emails(emails, options=None):
    """이메일 리스트로 Enrichment (정보 보강)"""
    return run_enrichment([{"email": e} for e in emails if e], options, True, "email_enriched")


def enrich_by_companies(companies, options=None):
    """회사+이름 조합으로 Enrichment"""
    return run_enrichment(companies, options, True, "company_enriched")


# ============================================================
# 실행
# ============================================================

if __name__ == "__main__":
    # ---- 방법 1: 직접 리스트 ----
    # leads = run_enrichment()

    # ---- 방법 2: CSV/Excel 파일 ----
    # leads = enrich_from_file("input/targets.csv")
    # leads = enrich_from_file("input/targets.xlsx", max_count=10)

    # ---- 방법 3: LinkedIn URL ----
    # leads = enrich_by_linkedin([
    #     "https://www.linkedin.com/in/person1",
    # ])

    # ---- 방법 4: 이메일 보강 ----
    # leads = enrich_by_emails(["ceo@company.com"])

    # ---- 방법 5: 회사+이름 ----
    # leads = enrich_by_companies([
    #     {"first_name": "Tim", "last_name": "Cook", "domain": "apple.com"},
    # ])

    leads = run_enrichment()
